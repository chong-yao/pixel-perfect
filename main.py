"""
Image-to-Excel and Excel-to-Image Converter (Optimized)

This script provides two main functionalities:
1.  Convert an image file (e.g., PNG, JPG) into a Microsoft Excel spreadsheet,
    representing the image by coloring individual cells.
2.  Convert an Excel spreadsheet with colored cells back into an image file.

Configure the settings in the '--- USER SETTINGS ---' section below and run the script.
"""

from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
import time
import sys

# ==============================================================================
# --- USER SETTINGS ---
# (Configure these variables to control the script's behavior)
# ==============================================================================

# -- Choose the operation:
# 'to_excel': Converts the input image to an Excel file.
# 'to_image': Converts the input Excel file back to an image.
# 'both':     Runs both conversions sequentially.
OPERATION_MODE = 'both'

# -- Define file paths:
# The source image to convert into pixel art.
INPUT_IMAGE_PATH = Path("image-1920x1080")
# The path where the generated Excel file will be saved.
EXCEL_PATH = Path("pixels.xlsx")
# The path where the final image, restored from Excel, will be saved.
FINAL_IMAGE_PATH = Path("z")

# -- Set processing parameters:
# The resolution for the pixel art. Set X (width) and Y (height) for the grid.
# e.g., (128, 64) for a wide image. Set both to the same value for a square grid.
GRID_SIZE_X = 5568
GRID_SIZE_Y = 3132

# The multiplier for scaling up the final output image from the Excel file.
# A higher value results in a larger image.
UPSCALE_FACTOR = 5

# ==============================================================================
# --- CORE FUNCTIONS ---
# (The main logic for the conversions resides here)
# ==============================================================================

def image_to_excel(image_path: Path, output_excel_path: Path, grid_size_x: int = 64, grid_size_y: int = 64) -> None:
    """
    Converts an image into an Excel sheet by coloring cells to represent pixels.

    Args:
        image_path (Path): Path to the input image.
        output_excel_path (Path): Path to save the output Excel file.
        grid_size_x (int): The width of the pixel grid.
        grid_size_y (int): The height of the pixel grid.
    """
    print("--- Starting: Image to Excel Conversion ---")
    try:
        # 1. Open and resize the image using the fast NEAREST resampling method.
        with Image.open(image_path) as img:
            pixel_art = img.resize((grid_size_x, grid_size_y), Image.Resampling.NEAREST)
            pixel_art = pixel_art.convert('RGB')

        # 2. Create a new Excel workbook.
        wb = Workbook()
        ws = wb.active
        ws.title = "Pixel Art"

        # 3. Cache fill colors to avoid creating duplicate style objects, improving performance.
        color_fills = {}
        print(f"Processing {grid_size_x}x{grid_size_y} image and writing to Excel...")
        
        # --- OPTIMIZATION START ---
        # Get all pixel data at once using getdata(), which is much faster than calling
        # getpixel() for each pixel individually inside a loop.
        pixels = list(pixel_art.getdata())
        width, _ = pixel_art.size

        # Iterate over the flat list of pixels.
        for i, (r, g, b) in enumerate(pixels):
            hex_color = f"{r:02x}{g:02x}{b:02x}"

            # Use the cached fill if available, otherwise create and cache a new one.
            if hex_color not in color_fills:
                color_fills[hex_color] = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            
            # Calculate the cell's row and column from the pixel's index 'i'.
            # Adding 1 because Excel sheets are 1-indexed.
            row = i // width + 1
            col = i % width + 1
            
            ws.cell(row=row, column=col).fill = color_fills[hex_color]
        # --- OPTIMIZATION END ---

        # 4. Adjust cell dimensions to be square-like for better visual representation.
        print("Adjusting cell sizes...")
        CELL_WIDTH = 3
        CELL_HEIGHT = 18
        for i in range(1, grid_size_x + 1):
            ws.column_dimensions[get_column_letter(i)].width = CELL_WIDTH
        for i in range(1, grid_size_y + 1):
            ws.row_dimensions[i].height = CELL_HEIGHT

        # 5. Save the workbook.
        wb.save(output_excel_path)
        print(f"\nSuccessfully created pixel art in '{output_excel_path}'")

    except FileNotFoundError:
        print(f"Error: The input image file '{image_path}' was not found.")
        sys.exit(1)
    except Exception as e:
        print(f"An error occurred during image_to_excel: {e}")
        sys.exit(1)

def excel_to_image(excel_path: Path, output_image_path: Path, scale: int = 8) -> None:
    """
    Converts an Excel sheet with colored cells back into an image file.

    Args:
        excel_path (Path): Path to the input Excel file.
        output_image_path (Path): Path to save the generated image file.
        scale (int): Factor by which to scale the final image size.
    """
    print("--- Starting: Excel to Image Conversion ---")
    try:
        # 1. Load the workbook in read-only mode for better performance. This is a key
        #    optimization when you only need to read data and styles.
        wb = load_workbook(filename=excel_path, read_only=True, data_only=True)
        ws = wb.active

        # 2. Get the dimensions of the pixel art from the sheet.
        width = ws.max_column
        height = ws.max_row
        
        # 3. Create a new blank image and get direct access to its pixel data.
        #    Using .load() is significantly faster than calling .putpixel() repeatedly.
        img = Image.new('RGB', (width, height), color='white')
        pixels = img.load()

        print(f"Reading Excel file with dimensions {width}x{height}...")
        # 4. Iterate through cells and set pixel data.
        #    NOTE: Accessing cell.fill is the main performance bottleneck here, as
        #    openpyxl has to parse the underlying XML style information for each cell.
        #    ws.iter_rows() is the most efficient way to do this with this library.
        for row_index, row in enumerate(ws.iter_rows()):
            for col_index, cell in enumerate(row):
                fill = cell.fill
                # Ensure the cell has a fill with a color defined.
                if fill and fill.start_color and fill.start_color.rgb:
                    # openpyxl color format is AARRGGBB, so we slice off the 'FF' alpha part.
                    hex_color = fill.start_color.rgb[2:]
                    # Convert hex string to an RGB tuple.
                    rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
                    pixels[col_index, row_index] = rgb
                else:
                    # Default to white if no fill is found.
                    pixels[col_index, row_index] = (255, 255, 255)

        # 5. Scale up the final image for better visibility without blurring (NEAREST resampling).
        final_size = (width * scale, height * scale)
        final_image = img.resize(final_size, Image.Resampling.NEAREST)
        
        # 6. Save the final image.
        final_image.save(output_image_path)
        print(f"\nSuccessfully converted Excel sheet to '{output_image_path}'")

    except FileNotFoundError:
        print(f"Error: The input Excel file '{excel_path}' was not found.")
        sys.exit(1)
    except Exception as e:
        print(f"An error occurred during excel_to_image: {e}")
        sys.exit(1)

# ==============================================================================
# --- SCRIPT EXECUTION ---
# ==============================================================================
def main():
    """
    Main function to run the selected operations.
    """
    start_time = time.time()
    
    if OPERATION_MODE == 'to_excel':
        image_to_excel(
            image_path=INPUT_IMAGE_PATH,
            output_excel_path=EXCEL_PATH,
            grid_size_x=GRID_SIZE_X,
            grid_size_y=GRID_SIZE_Y
        )
        
    elif OPERATION_MODE == 'to_image':
        excel_to_image(
            excel_path=EXCEL_PATH,
            output_image_path=FINAL_IMAGE_PATH,
            scale=UPSCALE_FACTOR
        )

    elif OPERATION_MODE == 'both':
        # Run both operations sequentially.
        image_to_excel(
            image_path=INPUT_IMAGE_PATH,
            output_excel_path=EXCEL_PATH,
            grid_size_x=GRID_SIZE_X,
            grid_size_y=GRID_SIZE_Y
        )
        print("\n" + "="*50 + "\n") # Visual separator
        excel_to_image(
            excel_path=EXCEL_PATH,
            output_image_path=FINAL_IMAGE_PATH,
            scale=UPSCALE_FACTOR
        )
        
    else:
        print(f"Error: Invalid OPERATION_MODE '{OPERATION_MODE}'. Please choose 'to_excel', 'to_image', or 'both'.")

    end_time = time.time()
    print(f"\nTotal execution time: {round(end_time - start_time, 2)} seconds.")

if __name__ == '__main__':
    main()