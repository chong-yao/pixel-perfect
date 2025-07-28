"""
Image-to-Excel and Excel-to-Image Converter

This script provides two main functionalities:
1.  Convert an image file (e.g., PNG, JPG) into a Microsoft Excel spreadsheet,
    representing the image by coloring individual cells.
2.  Convert an Excel spreadsheet with colored cells back into an image file.

Configure the settings in the '--- USER SETTINGS ---' section below and run the script.
"""

from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
import time
import sys

# ==============================================================================
# --- USER SETTINGS ---
# (Configure these variables to control the script's behavior)
# ==============================================================================

# -- Choose the operation:
# 'to_excel': Converts the input image to an Excel file.
# 'to_image': Converts the input Excel file back to an image.
# 'both':     Runs both conversions sequentially.
OPERATION_MODE = 'to_excel'

# -- Define file paths:
# The source image to convert into pixel art.
INPUT_IMAGE_PATH = Path("profile-picture-dall-e-2.png")
# The path where the generated Excel file will be saved.
EXCEL_PATH = Path("pixel-art.xlsx")
# The path where the final image, restored from Excel, will be saved.
FINAL_IMAGE_PATH = Path("pixel-art.png")

# -- Set processing parameters:
# The resolution for the pixel art (e.g., 64 means a 64x64 grid). Only applies when generating Excel file.
GRID_SIZE = 64
# The multiplier for scaling up the final output image from the Excel file.
# A higher value results in a larger image.
UPSCALE_FACTOR = 200

# ==============================================================================
# --- CORE FUNCTIONS ---
# (The main logic for the conversions resides here)
# ==============================================================================

def image_to_excel(image_path: Path, output_excel_path: Path, pixel_size: int = 64) -> None:
    """
    Converts an image into an Excel sheet by coloring cells to represent pixels.

    Args:
        image_path (Path): Path to the input image.
        output_excel_path (Path): Path to save the output Excel file.
        pixel_size (int): The width and height of the pixel grid.
    """
    print("--- Starting: Image to Excel Conversion ---")
    try:
        # 1. Open and resize the image using the fast NEAREST resampling method.
        with Image.open(image_path) as img:
            pixel_art = img.resize((pixel_size, pixel_size), Image.Resampling.NEAREST)
            pixel_art = pixel_art.convert('RGB')

        # 2. Create a new Excel workbook.
        wb = Workbook()
        ws = wb.active
        ws.title = "Pixel Art"

        # 3. Cache fill colors to avoid creating duplicate style objects, improving performance.
        color_fills = {}
        print(f"Processing {pixel_size}x{pixel_size} image and writing to Excel...")
        for row in range(pixel_art.height):
            for col in range(pixel_art.width):
                r, g, b = pixel_art.getpixel((col, row))
                hex_color = f"{r:02x}{g:02x}{b:02x}"

                if hex_color not in color_fills:
                    color_fills[hex_color] = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                
                ws.cell(row=row + 1, column=col + 1).fill = color_fills[hex_color]

        # 4. Adjust cell dimensions to be square-like for better visual representation.
        print("Adjusting cell sizes...")
        # Constants for cell dimensions
        CELL_WIDTH = 3
        CELL_HEIGHT = 18
        for i in range(1, pixel_size + 1):
            ws.column_dimensions[get_column_letter(i)].width = CELL_WIDTH
            ws.row_dimensions[i].height = CELL_HEIGHT

        # 5. Save the workbook.
        wb.save(output_excel_path)
        print(f"\nSuccessfully created pixel art in '{output_excel_path}'")

    except FileNotFoundError:
        print(f"Error: The input image file '{image_path}' was not found.")
        sys.exit(1)
    except Exception as e:
        print(f"An error occurred during image_to_excel: {e}")
        sys.exit(1)

def excel_to_image(excel_path: Path, output_image_path: Path, scale: int = 8) -> None:
    """
    Converts an Excel sheet with colored cells back into an image file.

    Args:
        excel_path (Path): Path to the input Excel file.
        output_image_path (Path): Path to save the generated image file.
        scale (int): Factor by which to scale the final image size.
    """
    print("--- Starting: Excel to Image Conversion ---")
    try:
        # 1. Load the workbook in read-only mode for better performance.
        wb = load_workbook(filename=excel_path, read_only=True, data_only=True)
        ws = wb.active

        # 2. Get the dimensions of the pixel art from the sheet.
        width = ws.max_column
        height = ws.max_row
        
        # 3. Create a new blank image.
        img = Image.new('RGB', (width, height), color='white')
        pixels = img.load() # Accessing pixels directly is faster.

        print(f"Reading Excel file with dimensions {width}x{height}...")
        # 4. Iterate through cells and set pixel data.
        for row_index, row in enumerate(ws.iter_rows()):
            for col_index, cell in enumerate(row):
                fill = cell.fill
                # Ensure the cell has a fill with a color defined.
                if fill and fill.start_color and fill.start_color.rgb:
                    # openpyxl color format is AARRGGBB, so we slice off the 'FF' alpha part.
                    hex_color = fill.start_color.rgb[2:]
                    rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
                    pixels[col_index, row_index] = rgb
                else:
                    pixels[col_index, row_index] = (255, 255, 255) # Default to white if no fill.

        # 5. Scale up the final image for better visibility without blurring.
        final_size = (width * scale, height * scale)
        final_image = img.resize(final_size, Image.Resampling.NEAREST)
        
        # 6. Save the final image.
        final_image.save(output_image_path)
        print(f"\nSuccessfully converted Excel sheet to '{output_image_path}'")

    except FileNotFoundError:
        print(f"Error: The input Excel file '{excel_path}' was not found.")
        sys.exit(1)
    except Exception as e:
        print(f"An error occurred during excel_to_image: {e}")
        sys.exit(1)

# ==============================================================================
# --- SCRIPT EXECUTION ---
# ==============================================================================
def main():
    """
    Main function to run the selected operations.
    """
    start_time = time.time()
    
    if OPERATION_MODE == 'to_excel':
        image_to_excel(
            image_path=INPUT_IMAGE_PATH,
            output_excel_path=EXCEL_PATH,
            pixel_size=GRID_SIZE
        )
        
    elif OPERATION_MODE == 'to_image':
        excel_to_image(
            excel_path=EXCEL_PATH,
            output_image_path=FINAL_IMAGE_PATH,
            scale=UPSCALE_FACTOR
        )

    elif OPERATION_MODE == 'both':
        # Run both operations sequentially.
        image_to_excel(
            image_path=INPUT_IMAGE_PATH,
            output_excel_path=EXCEL_PATH,
            pixel_size=GRID_SIZE
        )
        print("\n" + "="*50 + "\n") # Visual separator
        excel_to_image(
            excel_path=EXCEL_PATH,
            output_image_path=FINAL_IMAGE_PATH,
            scale=UPSCALE_FACTOR
        )
        
    else:
        print(f"Error: Invalid OPERATION_MODE '{OPERATION_MODE}'. Please choose 'to_excel', 'to_image', or 'both'.")

    end_time = time.time()
    print(f"\nTotal execution time: {end_time - start_time:.2f} seconds.")

if __name__ == '__main__':
    main()